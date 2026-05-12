#!/usr/bin/env python3

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - import guard
    raise SystemExit(
        "openpyxl is required to export the workbook queue. "
        "Use a Python runtime that includes openpyxl, such as the bundled Codex workspace runtime."
    ) from exc

SHEET_CONFIG = {
    "Backflow Article Bank": "backflow_article_bank",
    "Plumbing Water Article Bank": "plumbing_water_article_bank",
    "Local Regulation Article Bank": "local_regulation_article_bank",
    "Commercial Facility Bank": "commercial_facility_bank",
    "Refresh Localized Bank": "refresh_localized_bank",
}

WORKBOOK_HEADERS_ROW = 2
WORKBOOK_DATA_START_ROW = 3
DEFAULT_OUTPUT = Path("site/data/generated/blog-queue.json")


def slugify(value: str) -> str:
    return (
        re.sub(r"[^a-z0-9]+", "-", value.lower().replace("&", "and"))
        .strip("-")
    )


def split_csvish(value: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in re.split(r"[,\n;]+", value) if item.strip()]


def infer_geography(title: str) -> str:
    lowered = title.lower()
    geography_markers = [
        "california",
        "los angeles county",
        "orange county",
        "san diego county",
        "ventura county",
        "riverside county",
        "san bernardino county",
        "los angeles",
        "orange county",
        "san diego",
    ]
    for marker in geography_markers:
        if marker in lowered:
            return marker.title()
    return ""


def infer_template_type(sheet_name: str, title: str, pillar: str, primary_keyword: str) -> str:
    lowered = " ".join([sheet_name, title, pillar, primary_keyword]).lower()
    if sheet_name == "Commercial Facility Bank":
        return "commercial_facility_article"
    if sheet_name == "Local Regulation Article Bank":
        return "compliance_article"
    if sheet_name == "Refresh Localized Bank" and ("2026" in lowered or "refresh" in lowered):
        return "refresh_update_article"
    if infer_geography(title):
        return "localized_service_article"
    if any(term in lowered for term in ["cost", "wasting", "roi", "repair", "savings", "leak"]):
        return "cost_risk_article"
    if any(term in lowered for term in ["rule", "regulation", "requirement", "compliance", "annually"]):
        return "compliance_article"
    return "stat_explainer"


def infer_cta_type(sheet_name: str, audience: str) -> str:
    if sheet_name == "Commercial Facility Bank" or "property manager" in audience.lower():
        return "commercial_compliance_review"
    if sheet_name == "Plumbing Water Article Bank":
        return "request_plumbing_inspection"
    return "schedule_backflow_service"


def source_notes_for_rows(stats_sheet, row_numbers: list[int]) -> list[dict[str, object]]:
    source_stats: list[dict[str, object]] = []
    for row_number in row_numbers:
        row = [cell for cell in next(stats_sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True))]
        source_stats.append(
            {
                "repositoryRow": row_number,
                "dataPoint": row[2] or "",
                "value": row[3] or "",
                "scope": row[4] or "",
                "timePeriod": row[5] or "",
                "whyItMatters": row[6] or "",
                "sourceOrganization": row[7] or "",
                "sourceTitle": row[8] or "",
                "sourceUrl": row[9] or "",
                "authorityType": row[10] or "",
                "notes": row[11] if len(row) > 11 else None,
            }
        )
    return source_stats


def workbook_to_queue(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    stats_sheet = workbook["Stats Repository"]
    queue_records: list[dict[str, object]] = []

    for sheet_name in SHEET_CONFIG:
        worksheet = workbook[sheet_name]
        headers = [cell or "" for cell in next(
            worksheet.iter_rows(
                min_row=WORKBOOK_HEADERS_ROW,
                max_row=WORKBOOK_HEADERS_ROW,
                values_only=True,
            )
        )]
        for row in worksheet.iter_rows(min_row=WORKBOOK_DATA_START_ROW, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue
            data = dict(zip(headers, row))
            publish_order = int(data.get("Publish order") or len(queue_records) + 1)
            title = str(data.get("Working title") or "").strip()
            primary_keyword = str(data.get("Primary keyword / intent") or "").strip()
            audience = str(data.get("Best-fit audience") or "").strip()
            repository_rows = [int(item) for item in split_csvish(str(data.get("Repository row(s)") or "")) if item.isdigit()]
            slug = slugify(title)
            notes: list[str] = []
            template_type = infer_template_type(
                sheet_name,
                title,
                str(data.get("Pillar") or ""),
                primary_keyword,
            )
            if sheet_name in {"Local Regulation Article Bank", "Refresh Localized Bank"}:
                notes.append("Requires reviewer sign-off before publication.")
            geography = infer_geography(title)
            if geography:
                notes.append(f"Carry the {geography} angle through the hero, body, and related links.")

            queue_records.append(
                {
                    "id": f"{SHEET_CONFIG[sheet_name]}-{publish_order:03d}",
                    "articleTitle": title,
                    "category": str(data.get("Pillar") or "").strip(),
                    "sheetName": sheet_name,
                    "audience": audience,
                    "geography": geography,
                    "templateType": template_type,
                    "primaryKeyword": primary_keyword,
                    "secondaryKeywords": [],
                    "sourceStats": source_notes_for_rows(stats_sheet, repository_rows),
                    "ctaType": infer_cta_type(sheet_name, audience),
                    "slug": slug,
                    "status": "Queued",
                    "notes": " ".join(notes).strip(),
                }
            )

    return queue_records


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("Usage: export_blog_workbook_queue.py /absolute/path/to/workbook.xlsx [output.json]", file=sys.stderr)
        return 1

    workbook_path = Path(argv[1]).expanduser().resolve()
    output_path = (
        Path(argv[2]).expanduser().resolve() if len(argv) > 2 else DEFAULT_OUTPUT.resolve()
    )

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    queue_records = workbook_to_queue(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(f"{json.dumps(queue_records, indent=2)}\n", encoding="utf8")

    print(f"Wrote {len(queue_records)} queue records to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
